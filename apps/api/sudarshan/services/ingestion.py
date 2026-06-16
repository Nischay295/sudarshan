import csv
import io
from datetime import datetime
from decimal import Decimal
import openpyxl

def parse_ingestion_file(file_content: bytes, filename: str) -> list[dict]:
    """
    Parses a CSV or Excel (.xlsx) file and maps the columns to standard transaction payload fields.
    Returns a list of dictionaries ready to be mapped to ManualTransactionCreate.
    """
    rows = []
    
    if filename.lower().endswith(".xlsx"):
        # Parse XLSX using openpyxl
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        if not wb.sheetnames:
            raise ValueError("Excel file contains no worksheets.")
            
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            # Read header row
            headers = []
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col).value
                headers.append(str(val).strip() if val is not None else f"col_{col}")
                
            # Read data rows
            for row in range(2, sheet.max_row + 1):
                row_dict = {}
                has_data = False
                for col in range(1, len(headers) + 1):
                    val = sheet.cell(row=row, column=col).value
                    if val is not None:
                        has_data = True
                    row_dict[headers[col - 1]] = val
                if has_data:
                    rows.append(row_dict)
                
    else:
        # Parse CSV (fallback)
        try:
            text_content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text_content = file_content.decode("latin-1")
            
        csv_file = io.StringIO(text_content)
        reader = csv.DictReader(csv_file)
        for row in reader:
            if any(row.values()): # skip completely empty rows
                rows.append(dict(row))

    # Standardize row headers and convert fields
    standard_rows = []
    for row in rows:
        standardized = {}
        for k, v in row.items():
            clean_k = k.lower().replace(" ", "_").replace("-", "_").strip()
            standardized[clean_k] = v
            
        # Mapping helpers
        entry_date = _get_value_by_keys(standardized, ["entry_date", "date", "tx_date", "transaction_date"])
        description = _get_value_by_keys(standardized, ["description", "narration", "particulars", "desc", "details"])
        counterparty = _get_value_by_keys(standardized, ["counterparty", "party", "vendor", "customer", "payee"])
        gst_rate = _get_value_by_keys(standardized, ["gst_rate", "rate", "gst_percentage", "gst_%", "tax_rate"])
        gst_treatment = _get_value_by_keys(standardized, ["gst_treatment", "treatment", "tax_treatment"])
        payment_account = _get_value_by_keys(standardized, ["payment_account_code", "payment_account", "account_code", "account"])

        # Check for debit/credit columns first
        debit = _get_value_by_keys(standardized, ["debit", "withdrawal", "dr", "amount_paid", "paid_out"])
        credit = _get_value_by_keys(standardized, ["credit", "deposit", "cr", "amount_received", "received"])

        amount = None
        flow = None

        def _clean_num(val):
            if val is None or str(val).strip() == "" or str(val).lower() == "none":
                return None
            try:
                clean_val = str(val).replace(",", "").replace("'", "").replace('"', "").strip()
                if not clean_val:
                    return None
                return Decimal(clean_val)
            except Exception:
                return None

        debit_val = _clean_num(debit)
        credit_val = _clean_num(credit)

        if debit_val is not None and debit_val > Decimal("0.00"):
            amount = debit_val
            flow = "expense"
        elif credit_val is not None and credit_val > Decimal("0.00"):
            amount = credit_val
            flow = "income"
        else:
            amount_raw = _get_value_by_keys(standardized, ["amount", "value", "tx_amount", "base_amount", "taxable_amount"])
            amount = _clean_num(amount_raw)
            flow_raw = _get_value_by_keys(standardized, ["flow", "type", "flow_type", "transaction_type"])
            flow = str(flow_raw).lower().strip() if flow_raw else None

        # Resolve flow using description/details keyword matches
        if description:
            desc_str = str(description).lower().strip()
            if "sale" in desc_str or "revenue" in desc_str:
                flow = "sale"
            elif "purchase" in desc_str or "asset" in desc_str:
                flow = "purchase"
            elif "income" in desc_str:
                flow = "income"
            elif "expense" in desc_str or "rent" in desc_str or "salary" in desc_str or "travel" in desc_str or "fee" in desc_str:
                flow = "expense"
            elif "payment" in desc_str:
                flow = "payment"
            elif "receipt" in desc_str:
                flow = "receipt"
            elif "capital" in desc_str or "contribution" in desc_str:
                flow = "owner_contribution"
            elif "loan" in desc_str:
                flow = "loan_received"

        if not flow:
            flow = "expense"

        # Format and validate individual fields
        if not entry_date or not description or amount is None:
            continue # Skip row if mandatory fields are missing
            
        try:
            # Parse Date
            if isinstance(entry_date, datetime):
                parsed_date = entry_date.date()
            elif hasattr(entry_date, "date"): # handle date object from openpyxl
                parsed_date = entry_date
            else:
                parsed_date = _parse_date_string(str(entry_date).strip())
                
            # Parse Amount
            parsed_amount = Decimal(str(amount).replace(",", "").strip()).quantize(Decimal("0.01"))
            
            # Parse GST Rate
            parsed_gst_rate = Decimal("0.00")
            if gst_rate is not None:
                parsed_gst_rate = Decimal(str(gst_rate).replace("%", "").strip()).quantize(Decimal("0.01"))
                
            # Parse Flow
            parsed_flow = str(flow).lower().strip() if flow else "expense"
            if parsed_flow not in ["sale", "purchase", "expense", "income", "receipt", "payment", "owner_contribution", "loan_received"]:
                parsed_flow = "expense"
                
            # Parse GST Treatment
            parsed_treatment = str(gst_treatment).lower().strip() if gst_treatment else "none"
            if parsed_treatment not in ["none", "exempt", "intra_state", "inter_state"]:
                parsed_treatment = "none"
                
            standard_rows.append({
                "entry_date": parsed_date.strftime("%Y-%m-%d"),
                "description": str(description).strip(),
                "amount": str(parsed_amount),
                "flow": parsed_flow,
                "counterparty": str(counterparty).strip() if counterparty else None,
                "gst_rate": str(parsed_gst_rate),
                "gst_treatment": parsed_treatment,
                "payment_account_code": str(payment_account).strip() if payment_account else "1010"
            })
        except Exception:
            # Skip invalid rows silently or let the ingestion endpoint report exceptions
            continue
            
    return standard_rows

def _get_value_by_keys(d: dict, keys: list[str]):
    for k in keys:
        if k in d:
            return d[k]
    return None

def _parse_date_string(date_str: str) -> datetime:
    formats = ["%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y", "%b %d, %Y", "%d %b %Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass
    raise ValueError(f"Could not parse date string: {date_str}")
